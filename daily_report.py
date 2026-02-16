#!/usr/bin/env python3
"""
IT Daily Report Generator
Generates: IT_Daily_Report.xlsx with 4 tabs
"""

import pandas as pd

# All store names for tracking
ALL_STORES = [
    'Jenkins', 'Neon', 'Harlan 1', 'Harlan 2', 'Hyden', 'PMM', 'Isom', 
    'Whitesburg', 'Hazard 2', 'Ermine', 'Hindman 2', 
    'Hindman 1', 'Martin', 'Jackson', 'Hazard 3', 'Dryfork', 'Pound', 'Catnip (Nicholasville)', 'Marrowbone', 'Elkhorn City', 'Chloe', 
    'Caney', 'Belfrey', 'Phelps', 'Virgie', 'Harold', 'Allen', 'Goody', 
    'Zebulon', 'Pikeville', 'South', 'North', 'Prestonsburg 1', 'Ivel', 
    'Justiceville', 'Salyersville', 'Grundy', 'West Liberty', 
    'Prestonsburg 2', 'Prestonsburg 3'
]

def load_and_process_data():
    """Load and process ticket data"""
    print("\nLoading data...")
    active = pd.read_csv('active_tickets.csv')
    closed = pd.read_csv('closed_tickets.csv')
    
    # Filter to IT Helpdesk team only
    if 'Team' in active.columns:
        active = active[active['Team'] == 'IT Helpdesk']
    if 'Team' in closed.columns:
        closed = closed[closed['Team'] == 'IT Helpdesk']
    
    print(f"✓ Filtered to IT Helpdesk team only")
    
    # Filter closed to last 7 days using Last Modified Date
    closed['Last Modified Date'] = pd.to_datetime(closed['Last Modified Date'], errors='coerce', utc=True)
    seven_days_ago = pd.Timestamp.now(tz='UTC') - pd.Timedelta(days=7)
    closed = closed[closed['Last Modified Date'] >= seven_days_ago]
    
    # Remove timezone info from all datetime columns to avoid Excel issues
    for col in closed.columns:
        if pd.api.types.is_datetime64_any_dtype(closed[col]):
            closed[col] = closed[col].dt.tz_localize(None)
    
    for col in active.columns:
        if pd.api.types.is_datetime64_any_dtype(active[col]):
            active[col] = active[col].dt.tz_localize(None)
    
    # Treat "Resolved" status as closed too - pull from active
    active_copy = active.copy()
    resolved = active_copy[active_copy['Status'].str.lower().str.contains('resolved', na=False)]
    active = active_copy[~active_copy['Status'].str.lower().str.contains('resolved', na=False)]
    
    # Combine resolved with closed
    if len(resolved) > 0:
        closed = pd.concat([closed, resolved], ignore_index=True)
    
    # Add Status_Type BEFORE combining
    active['Status_Type'] = 'Active'
    closed['Status_Type'] = 'Closed'
    
    print(f"✓ Active tickets: {len(active)}")
    print(f"✓ Closed/Resolved tickets (last 7 days): {len(closed)}")
    
    return active, closed

def generate_it_daily_report(active, closed):
    """Generate IT Daily Report with 3 tabs"""
    print("\n" + "=" * 50)
    print("IT DAILY REPORT")
    print("=" * 50)
    
    filename = 'IT_Daily_Report.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # ========== TAB 1: ASSIGNEE SUMMARY ==========
    print("\n📊 Creating Tab 1: Assignee Summary...")
    
    all_for_summary = pd.concat([active, closed], ignore_index=True)
    assigned = all_for_summary[all_for_summary['Assignee'].notna() & (all_for_summary['Assignee'] != '')]
    
    summary_data = []
    for assignee in sorted(assigned['Assignee'].unique()):
        assignee_tickets = assigned[assigned['Assignee'] == assignee]
        active_count = len(assignee_tickets[assignee_tickets['Status_Type'] == 'Active'])
        closed_count = len(assignee_tickets[assignee_tickets['Status_Type'] == 'Closed'])
        
        summary_data.append({
            'Assignee': assignee,
            'Active Tickets': active_count,
            'Closed (Last 7 Days)': closed_count,
            'Total': active_count + closed_count
        })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Assignee Summary', index=False)
    
    # ========== TAB 2: ACTIVE TICKETS (SEPARATED BY ASSIGNEE) ==========
    print("📋 Creating Tab 2: Active Tickets...")
    
    # Start building the worksheet manually with spacing between assignees
    active_sheet_data = []
    
    # Add unassigned section first
    unassigned = active[active['Assignee'].isna() | (active['Assignee'] == '')]
    if len(unassigned) > 0:
        active_sheet_data.append(['UNASSIGNED', '', '', '', '', '', ''])
        active_sheet_data.append(['Ticket ID', 'Request', 'Requester', 'Priority', 'Category', 'Created Date', 'Assignee'])
        for _, ticket in unassigned.iterrows():
            # Convert date to string to avoid timezone issues
            created_date = str(ticket.get('Created Date', ''))[:19] if pd.notna(ticket.get('Created Date', '')) else ''
            
            active_sheet_data.append([
                ticket.get('Id', ''),
                ticket.get('Request', ''),
                ticket.get('Requester', ''),
                ticket.get('Priority', ''),
                ticket.get('Category', ''),
                created_date,
                'UNASSIGNED'
            ])
        active_sheet_data.append(['', '', '', '', '', '', ''])  # Blank row
    
    # Add each assignee's section
    assigned_active = active[active['Assignee'].notna() & (active['Assignee'] != '')]
    for assignee in sorted(assigned_active['Assignee'].unique()):
        assignee_tickets = assigned_active[assigned_active['Assignee'] == assignee]
        
        active_sheet_data.append([assignee, '', '', '', '', '', ''])
        active_sheet_data.append(['Ticket ID', 'Request', 'Requester', 'Priority', 'Category', 'Created Date', 'Assignee'])
        
        for _, ticket in assignee_tickets.iterrows():
            # Convert date to string to avoid timezone issues
            created_date = str(ticket.get('Created Date', ''))[:19] if pd.notna(ticket.get('Created Date', '')) else ''
            
            active_sheet_data.append([
                ticket.get('Id', ''),
                ticket.get('Request', ''),
                ticket.get('Requester', ''),
                ticket.get('Priority', ''),
                ticket.get('Category', ''),
                created_date,
                assignee
            ])
        active_sheet_data.append(['', '', '', '', '', '', ''])  # Blank row
    
    # Write to Excel
    active_df = pd.DataFrame(active_sheet_data)
    active_df.to_excel(writer, sheet_name='Active Tickets', index=False, header=False)
    
    # Apply yellow highlight to unassigned section
    workbook = writer.book
    worksheet = writer.sheets['Active Tickets']
    
    from openpyxl.styles import PatternFill, Font, Alignment
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 10  # Ticket ID
    worksheet.column_dimensions['B'].width = 60  # Request (wider for wrapping)
    worksheet.column_dimensions['C'].width = 20  # Requester
    worksheet.column_dimensions['D'].width = 12  # Priority
    worksheet.column_dimensions['E'].width = 20  # Category
    worksheet.column_dimensions['F'].width = 20  # Created Date
    worksheet.column_dimensions['G'].width = 18  # Assignee
    
    # Apply formatting to all cells
    for row in range(1, len(active_sheet_data) + 1):
        # First pass: apply alignment to all cells
        for col in range(1, 8):
            cell = worksheet.cell(row, col)
            # Column B (Request) gets wrap text, others don't
            if col == 2:  # Request column
                cell.alignment = wrap_align
            else:
                cell.alignment = top_left_align
        
        # Second pass: apply highlighting and other formatting
        cell_value = worksheet.cell(row, 7).value  # Column G (Assignee)
        if cell_value == 'UNASSIGNED' or (row > 1 and worksheet.cell(row, 1).value == 'UNASSIGNED'):
            for col in range(1, 8):
                cell = worksheet.cell(row, col)
                cell.fill = yellow_fill
                # Reapply alignment after fill
                if col == 2:
                    cell.alignment = wrap_align
                else:
                    cell.alignment = top_left_align
        
        # Bold assignee name headers
        if row > 1 and worksheet.cell(row + 1, 1).value == 'Ticket ID':
            worksheet.cell(row, 1).font = bold_font
    
    # ========== TAB 3: CLOSED TICKETS (SEPARATED BY ASSIGNEE) ==========
    print("✅ Creating Tab 3: Closed Tickets...")
    
    closed_sheet_data = []
    
    assigned_closed = closed[closed['Assignee'].notna() & (closed['Assignee'] != '')]
    
    if len(assigned_closed) == 0:
        closed_sheet_data.append(['No closed tickets in the last 7 days'])
    else:
        for assignee in sorted(assigned_closed['Assignee'].unique()):
            assignee_tickets = assigned_closed[assigned_closed['Assignee'] == assignee]
            
            closed_sheet_data.append([assignee, '', '', '', '', '', '', ''])
            closed_sheet_data.append(['Ticket ID', 'Request', 'Requester', 'Priority', 'Category', 'Created Date', 'Last Modified Date', 'Assignee'])
            
            for _, ticket in assignee_tickets.iterrows():
                # Convert dates to strings to avoid timezone issues
                created_date = str(ticket.get('Created Date', ''))[:19] if pd.notna(ticket.get('Created Date', '')) else ''
                modified_date = str(ticket.get('Last Modified Date', ''))[:19] if pd.notna(ticket.get('Last Modified Date', '')) else ''
                
                closed_sheet_data.append([
                    ticket.get('Id', ''),
                    ticket.get('Request', ''),
                    ticket.get('Requester', ''),
                    ticket.get('Priority', ''),
                    ticket.get('Category', ''),
                    created_date,
                    modified_date,
                    assignee
                ])
            closed_sheet_data.append(['', '', '', '', '', '', '', ''])  # Blank row
    
    closed_df = pd.DataFrame(closed_sheet_data)
    closed_df.to_excel(writer, sheet_name='Closed Tickets', index=False, header=False)
    
    # Apply formatting
    worksheet_closed = writer.sheets['Closed Tickets']
    
    from openpyxl.styles import Alignment
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Set column widths
    worksheet_closed.column_dimensions['A'].width = 10  # Ticket ID
    worksheet_closed.column_dimensions['B'].width = 60  # Request (wider for wrapping)
    worksheet_closed.column_dimensions['C'].width = 20  # Requester
    worksheet_closed.column_dimensions['D'].width = 12  # Priority
    worksheet_closed.column_dimensions['E'].width = 20  # Category
    worksheet_closed.column_dimensions['F'].width = 20  # Created Date
    worksheet_closed.column_dimensions['G'].width = 20  # Last Modified Date
    worksheet_closed.column_dimensions['H'].width = 18  # Assignee
    
    # Apply formatting to all cells
    for row in range(1, len(closed_sheet_data) + 1):
        for col in range(1, 9):
            cell = worksheet_closed.cell(row, col)
            # Column B (Request) gets wrap text, others don't
            if col == 2:  # Request column
                cell.alignment = wrap_align
            else:
                cell.alignment = top_left_align
        
        # Bold assignee name headers
        if row > 1 and worksheet_closed.cell(row + 1, 1).value == 'Ticket ID':
            worksheet_closed.cell(row, 1).font = bold_font
    
    # ========== TAB 4: STORE COUNTS ==========
    print("🏪 Creating Tab 4: Store Counts...")
    
    # Combine active and closed for total counts
    all_for_stores = pd.concat([active, closed], ignore_index=True)
    
    # Build store counts
    store_count_data = []
    for store in ALL_STORES:
        # Count tickets where Requester contains the store name
        count = len(all_for_stores[all_for_stores['Requester'].str.contains(store, case=False, na=False)])
        store_count_data.append({
            'Store Name': store,
            'Number of Tickets': count
        })
    
    # Sort by count descending
    store_df = pd.DataFrame(store_count_data)
    store_df = store_df.sort_values('Number of Tickets', ascending=False)
    store_df.to_excel(writer, sheet_name='Store Counts', index=False)
    
    # Apply formatting to Store Counts tab
    worksheet_stores = writer.sheets['Store Counts']
    
    # Set column widths
    worksheet_stores.column_dimensions['A'].width = 25  # Store Name
    worksheet_stores.column_dimensions['B'].width = 18  # Number of Tickets
    
    # Apply top-left alignment to all cells
    for row in range(1, len(store_df) + 2):  # +2 for header row
        for col in range(1, 3):
            cell = worksheet_stores.cell(row, col)
            cell.alignment = top_left_align
    
    # Bold the header row
    for col in range(1, 3):
        worksheet_stores.cell(1, col).font = bold_font
    
    writer.close()
    print(f"✓ Saved: {filename}")

def main():
    print("=" * 50)
    print("IT DAILY REPORT GENERATOR")
    print("=" * 50)
    
    # Load and process data
    active, closed = load_and_process_data()
    
    # Generate report
    generate_it_daily_report(active, closed)
    
    print("\n" + "=" * 50)
    print("✅ REPORT GENERATED!")
    print("=" * 50)
    print("\nFile created: IT_Daily_Report.xlsx")
    print("\n")

if __name__ == "__main__":
    main()
